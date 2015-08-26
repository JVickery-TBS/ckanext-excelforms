import openpyxl

from ckanext.recombinant.plugins import get_table
from ckanext.recombinant.errors import RecombinantException
from ckanext.recombinant.datatypes import data_store_type

def xls_template(dataset_type, org):
    """
    return an openpyxl.Workbook object containing the sheet and header fields
    for passed dataset_type and org.
    """
    t = get_table(dataset_type)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = t['xls_sheet_name']
    for n, key in enumerate(t['xls_organization_info']):
        for e in org['extras']:
            if e['key'] == key:
                sheet.cell(row=1, column=n + 1).value = e['value']
                break
        else:
            sheet.cell(row=1, column=n + 1).value = org.get(key, '')
    apply_styles(t['xls_organization_style'], sheet.row_dimensions[1])

    for n, field in enumerate(t['fields']):
        sheet.cell(row=2, column=n + 1).value = field['label']
        # jumping through openpyxl hoops:
        col = sheet.column_dimensions[openpyxl.cell.get_column_letter(n + 1)]
        col.width = field['xls_column_width']
        # FIXME: format only below header
        col.number_format = data_store_type[field['datastore_type']].xl_format
    apply_styles(t['xls_header_style'], sheet.row_dimensions[2])

    sheet.freeze_panes = sheet['A3']
    return book


def apply_styles(config, target):
    """
    apply styles from config to target

    currently supports PatternFill and Font
    """
    pattern_fill = config.get('PatternFill')
    if pattern_fill:
        target.fill = openpyxl.styles.PatternFill(**pattern_fill)
    font = config.get('Font')
    if font:
        target.font = openpyxl.styles.Font(**font)

